#reading a excel file and printing number of rows in the workbook and priting all columns in the workbook
import openpyxl as xl

#wb = xl.load_workbook("C:\\Users\\ajayp\\OneDrive\\Desktop\\Python\\student_data.xlsx") #reading Excel file
wb = xl.load_workbook("C:\\Users\\d862757\\Downloads\\archive (2)\\student_data.xlsx") #reading Excel file
print(wb.sheetnames) #to get names of all the sheets
sheet1 = wb.worksheets[0] #first work sheet
print(sheet1.max_row) #gives  row count
print(sheet1.max_column) #gives coloumn count
print(sheet1.min_column)

#looping

#for s in sheet.min_column:
    #print(s)

#sheet.merge_cells("A1:B2") - merges cells

#returns column names
#for s in sheet.iter_cols(sheet.min_column,sheet.max_column):
    #print(s[0].value)

#retriving column values
li = []
for i in list(sheet1.columns)[11]:
    li.append(i.value)

li.pop(0)
print(li) 

#creating a new sheet
sheet2 = wb.create_sheet("sheet2") #inserts at last position

wb.save(filename = "C:\\Users\\d862757\\Downloads\\archive (2)\\student_data_update.xlsx")